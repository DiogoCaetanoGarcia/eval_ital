# pyinstaller --onefile converte_pasta_xml_em_xlsx.py

from xml.dom import minidom
from os import listdir, path
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from PyQt5.QtWidgets import QApplication, QWidget, QFileDialog
import time
import sys
import re
import os
from collections import Counter
import matplotlib.pyplot as plt
from PIL import Image

def find_fields(xml_file_name, elms_attrs):
	xmldoc = minidom.parse(xml_file_name) # Read XML file
	cur_list = []
	for elms in elms_attrs:
		itemlist = xmldoc.getElementsByTagName(elms['tag']) # Look for tag
		if elms['count']: # We only want a count of this tag
			cur_list.append(len(itemlist))
		else:
			if (len(itemlist)==0): # Tag not found
				cur_list.append('')
			else: # Tag found
				attrs = [i.getAttribute(cur_attr) for i in itemlist for cur_attr in elms['attr']]
				if elms['get_unq']:
					attrs = list(set(attrs))
				s = '/'
				if elms['fnd_blnk']:
					l = len(elms['attr'])
					try:
						cur_index = attrs[0:-1:l].index('')
					except ValueError:
						cur_index = -1
					if cur_index!=-1:
						cur_index *= l
						cur_list.append(s.join(attrs[cur_index+1:cur_index+l]))
					else:
						cur_list.append('')
				else:
					cur_list.append(s.join(attrs))
	return cur_list

def simple_hist(cur_list, calc_perc=True):
	cur_dict = dict(Counter(cur_list))
	if '' in cur_dict.keys():
		cur_dict["Não-informado"] = cur_dict.pop('')
	if calc_perc:
		return [[dk, cur_dict[dk], cur_dict[dk]/len(cur_list)] for dk in cur_dict.keys()]
	else:
		return [[dk, cur_dict[dk]] for dk in cur_dict.keys()]

def f_remove_accents(old):
	new = old.lower()
	new = re.sub(r'[àáâãäå]', 'a', new)
	new = re.sub(r'[èéêë]', 'e', new)
	new = re.sub(r'[ìíîï]', 'i', new)
	new = re.sub(r'[òóôõöð]', 'o', new)
	new = re.sub(r'[ùúûü]', 'u', new)
	new = re.sub(r'[š]', 's', new)
	new = re.sub(r'[ž]', 'z', new)
	new = re.sub(r'[ýÿ]', 'y', new)
	new = re.sub(r'[ñ]', 'n', new)
	return new

def count_prod(prod_vals):
	prod_cnt = [sum([int(cur_p) for cur_p in prod_vals])]
	prod_cnt.append(prod_cnt[0]/len(prod_vals))
	return prod_cnt

def count_areas(areas_vals):
	full_count = len(areas_vals)
	areas_keys = set("/".join(areas_vals).split("/"))
	areas = []
	for ak in areas_keys:
		if len(ak)>0:
			cur_area = [ak,sum([ak in g for g in areas_vals])]
		else:
			cur_area = ["Não-informado",sum([len(g)==0 for g in areas_vals])]
		cur_area.append(cur_area[1]/full_count)
		areas.append(cur_area)
	areas.sort(key=lambda sort_key: sort_key[1], reverse=True)
	return areas

def count_prods(prods_vals, prods_keys=[]):
	full_count = len(prods_vals)
	if len(prods_keys)==0:
		prods_keys = set("/".join(prods_vals).split("/"))
	prods = []
	for ak in prods_keys:
		if len(prods)==0:
			prods = [ak in g for g in prods_vals]
		else:
			prods = [a or b for a,b in zip(prods, [ak in g for g in prods_vals])]
	return sum(prods)

def write_cells(sheet, sheet_header, sheet_data, col_num=1, row_num=1):
	for c,vh in enumerate(sheet_header):
		sheet.cell(row=row_num,column=col_num+c).value=vh
	for r,vl in enumerate(sheet_data):
		for c,v in enumerate(vl):
			sheet.cell(row=row_num+r+1,column=col_num+c).value=v

def func(pct, allvalues):
	absolute = int(pct / 100.*sum(allvalues))
	return "{:.1f}%\n({:d})".format(pct, absolute)

def write_cells_chart(cur_workbook, sheet_name, data_header, data_vals, col_num, row_num, chart_type, chart_title, chart_axis_titles, chart_sheetname, output_fig_dir):
	# Write data to sheet
	sheet = cur_workbook[sheet_name]
	write_cells(sheet, data_header, data_vals, col_num, row_num)
	if chart_type=='BarChart':
		cur_labels = [d[0] for d in data_vals]
		cur_data   = [d[1] for d in data_vals]
		fig = plt.figure() #figsize=[15, 5])
		plt.barh(cur_labels, cur_data)
		plt.xlabel(chart_axis_titles[0])
		plt.ylabel(chart_axis_titles[1])
		plt.grid(True)
		plt.title(chart_title)
		plt.tight_layout()
		# plt.show()
		fig.savefig(output_fig_dir + chart_sheetname + '.png')
		chart = BarChart()
		chart.x_axis.title = chart_axis_titles[0]
		chart.y_axis.title = chart_axis_titles[1]
	if chart_type=='PieChart':
		cur_labels = [d[0] for d in data_vals]
		cur_data   = [d[1] for d in data_vals]
		fig = plt.figure()
		plt.pie(cur_data, labels=cur_labels,
			autopct = lambda pct: func(pct, cur_data))
		plt.title(chart_title)
		# plt.show()
		fig.savefig(output_fig_dir + chart_sheetname + '.png')
		chart = PieChart()
	chart.type = 'bar'
	chart.title = chart_title
	data = Reference(sheet, min_col=col_num+1, max_col=col_num+len(data_header)-1,
		min_row=row_num,   max_row=row_num+len(data_vals))
	cats = Reference(sheet, min_col=col_num,   max_col=col_num,
		min_row=row_num+1, max_row=row_num+len(data_vals))
	chart.add_data(data, titles_from_data=True)
	chart.set_categories(cats)
	cs = cur_workbook.create_chartsheet(chart_sheetname)
	cs.add_chart(chart)

def xmls_2_xlsx(xml_file_folder, elms_attrs, output_file_name):
	# Criação do arquivo Excel
	wb = Workbook()
	ws = wb.create_sheet('Dados')
	del wb['Sheet']
	ws = wb['Dados']
	ws.append([e['fld_name'] for e in elms_attrs])
	cont = 0
	t0 = time.perf_counter()
	cur_list = []
	# Salvando resultados de cada arquivo XML no Excel
	for arq in listdir(xml_file_folder):
		if arq[-4:] == '.xml':
			cont += 1
			# if cont<=100:
			cur_list.append(find_fields(xml_file_folder+arq, elms_attrs))
			ws.append(cur_list[-1])
			if cont % 100 == 0:
				t1 = time.perf_counter() 
				print("%d: %s (%f s)" % (cont,arq,t1-t0))
				t0 = t1

	# Criar análises e gráficos
	wb.create_sheet('Análises')
	ws = wb['Análises']
	output_fig_dir = os.path.dirname(output_file_name) + "/"
	imgs_matplotlib = []

	# Contagem de brasileiros com doutorados, mestrados etc. na Itália
	brasileiros = [cl for cl in cur_list if cl[1]=="Brasil"]
	N_brasileiros = len(brasileiros)
	# Campos XML a serem lidos
	formacao_lbls = ["Doutorado", "Mestrado", "Especialização","Graduação"]
	# Palavras para procurar nestes campos
	palavras_italianas = ["universita ", "accademia ", "associazione ", "internazionale ", "ricerche ", "istituto ", "studi "]
	formacao_br = []
	# Para cada campo XML
	for k in range(len(formacao_lbls)):
		# Vamos remover todos acentos e deixar as letras minúsculas, para facilitar a busca
		cur_f = [f_remove_accents(p[k+10]) for p in brasileiros]
		# Vamos procurar as palavras no campo atual
		cur_formacao = [formacao_lbls[k],
			count_prods(cur_f, palavras_italianas)]
		cur_formacao.append(cur_formacao[1]/N_brasileiros)
		formacao_br.append(cur_formacao)
	# Vamos ordenar os valores encontrados, e salva-los no arquivo Excel e em um gráfico PNG
	formacao_br.sort(key=lambda sort_key: sort_key[1], reverse=True)
	write_cells_chart(wb, 'Análises', ["Formação de brasileiros na Itália", "Contagem"], [[f[0], f[1]] for f in formacao_br], 26, 1,
		"BarChart", 'Formação dos %d brasileiros na Itália' % N_brasileiros, ['Contagem', ''], "Formação_brasileiros", output_fig_dir)
	imgs_matplotlib.append("Formação_brasileiros")

	# Contagem de países listados no XML, salvando os resultados no Excel
	paises = simple_hist([cl[1] for cl in cur_list], False)
	write_cells(ws, ["País de nascimento", "Contagem"], paises, 3)

	# Separação dos italianos do XML, indicando o total de italianos no Excel
	italianos = [cl for cl in cur_list if cl[1]=="Itália"]
	N_italianos = len(italianos)
	write_cells(ws, ["Total de italianos"], [[len(italianos)]], 1)
	
	# Contagem de italianos por estado brasileiro, salvando resultados no arquivo Excel e em um gráfico PNG
	italianos_estado = simple_hist([cl[8] for cl in italianos], False)
	write_cells_chart(wb, 'Análises', ["Estado", "Contagem"], italianos_estado,
		6, 1, "PieChart", 'Italianos por estado', ['', 'Contagem'], "Italianos_estado", output_fig_dir)
	imgs_matplotlib.append("Italianos_estado")

	# Contagem ordenada de produções dos italianos
	prod_lbls = ["Trabalhos em eventos", "Artigos publicados", "Livros e capítulos", "Participação em projetos", "Patentes", "Processos ou técnicas", "Trabalho técnico", "Orientações (doutorado)", "Orientações (mestrado)", "Orientações (outras)"]
	prods = []
	for k in range(len(prod_lbls)):
		cur_prod = count_prod([p[k+18] for p in italianos])
		prods.append([prod_lbls[k],cur_prod[0], cur_prod[1]])
	prods.sort(key=lambda sort_key: sort_key[1], reverse=True)
	write_cells_chart(wb, 'Análises', ["Produção", "Contagem absoluta"], [[p[0], p[1]] for p in prods], 9, 1,
		"BarChart", 'Produção', ['Contagem', ''], "Produção", output_fig_dir)
	write_cells_chart(wb, 'Análises', ["Produção", "Contagem relativa"], [[p[0], p[2]] for p in prods], 11, 1,
		"BarChart", 'Produção/pessoa', ['Contagem', ''], "Produção_relativa", output_fig_dir)
	imgs_matplotlib.append("Produção")
	imgs_matplotlib.append("Produção_relativa")

	# Contagem de italianos com doutorados, mestrados etc
	formacao_lbls = ["Doutorado", "Mestrado", "Especialização","Graduação"]
	formacao_it = []
	for k in range(len(formacao_lbls)):
		cur_formacao = [formacao_lbls[k],sum([len(p[k+10])>0 for p in italianos])]
		cur_formacao.append(cur_formacao[1]/N_italianos)
		formacao_it.append(cur_formacao)
	formacao_it.sort(key=lambda sort_key: sort_key[1], reverse=True)
	write_cells_chart(wb, 'Análises', ["Formação dos italianos", "Contagem"], [[f[0], f[1]] for f in formacao_it], 23, 1,
		"BarChart", 'Formação dos %d italianos' % N_italianos, ['Contagem', ''], "Formação_italianos", output_fig_dir)
	imgs_matplotlib.append("Formação_italianos")

	# Contagem ordenada de cada area para os italianos
	areas_keys = ["Grande área de atuação", "Área de atuação", "Sub-área de atuação", "Especialidade"]
	areas_cnts = []
	for k in range(len(areas_keys)):
		areas_cnts.append(count_areas([cl[k+14] for cl in italianos]))
	for ai,ak in enumerate(areas_keys):
		imgs_matplotlib.append(ak.replace(" ", "_"))
		write_cells_chart(wb, 'Análises', [ak, "Contagem"], [[ac[0],ac[1]] for ac in areas_cnts[ai]], 14+ai*2, 1,
			"BarChart", ak, ['Contagem', ''], imgs_matplotlib[-1], output_fig_dir)	

	# União dos gráficos gerados em um PDF
	imgs_pil = []
	imgs_pil_list = []
	for k in range(len(imgs_matplotlib)):
		imgs_pil.append(Image.open(output_fig_dir + imgs_matplotlib[k] + ".png"))
		if k>0:
			imgs_pil_list.append(imgs_pil[-1].convert("RGB"))
		else:
			img1 = imgs_pil[-1].convert("RGB")
	img1.save(os.path.splitext(output_file_name)[0] + ".pdf", save_all=True, append_images=imgs_pil_list)

	# Salvando arquivo Excel
	wb.save(output_file_name)
	return

class dialogo(QWidget):
	def __init__(self, parent=None):
		QWidget.__init__(self, parent)
		self.folder = "data/2021_04/" # folder = "data/"
		self.output_file_name = "xml_to_excel.xlsx"
		dialog = QFileDialog()
		self.folder = dialog.getExistingDirectory(self, 'Escolha a pasta')
		self.folder = self.folder + '/'
		self.output_file_name, _ = dialog.getSaveFileName(self, 'Escolha o nome do arquivo de saída', "", "Excel files (*.xlsx)")

xml_fields = [{'fld_name':'Nome', 'tag':'DADOS-GERAIS', 'attr':['NOME-COMPLETO'], 'count':False, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'País de nascimento', 'tag':'DADOS-GERAIS', 'attr':['PAIS-DE-NASCIMENTO'], 'count':False, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Data de atualização', 'tag':'CURRICULO-VITAE', 'attr':['DATA-ATUALIZACAO'], 'count':False, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Id Lattes', 'tag':'CURRICULO-VITAE', 'attr':['NUMERO-IDENTIFICADOR'] , 'count':False, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Instituição de atuação', 'tag':'ENDERECO-PROFISSIONAL', 'attr':['NOME-INSTITUICAO-EMPRESA','NOME-ORGAO','NOME-UNIDADE'], 'count':False, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Vínculo', 'tag':'VINCULOS', 'attr':['ANO-FIM','OUTRO-ENQUADRAMENTO-FUNCIONAL-INFORMADO','OUTRO-VINCULO-INFORMADO','TIPO-DE-VINCULO'], 'count':False, 'fnd_blnk':True, 'get_unq':False},
	{'fld_name':'País de atuação', 'tag':'ENDERECO-PROFISSIONAL', 'attr':['PAIS'], 'count':False, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Cidade', 'tag':'ENDERECO-PROFISSIONAL', 'attr':['CIDADE'], 'count':False, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Estado', 'tag':'ENDERECO-PROFISSIONAL', 'attr':['UF'], 'count':False, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'CEP', 'tag':'ENDERECO-PROFISSIONAL', 'attr':['CEP'], 'count':False, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Doutorado', 'tag':'DOUTORADO', 'attr':['NOME-INSTITUICAO','CODIGO-INSTITUICAO','ANO-DE-CONCLUSAO','ANO-DE-OBTENCAO-DO-TITULO'], 'count':False, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Mestrado', 'tag':'MESTRADO', 'attr':['NOME-INSTITUICAO','CODIGO-INSTITUICAO','ANO-DE-CONCLUSAO','ANO-DE-OBTENCAO-DO-TITULO'], 'count':False, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Especialização', 'tag':'ESPECIALIZACAO', 'attr':['NOME-INSTITUICAO','CODIGO-INSTITUICAO','ANO-DE-CONCLUSAO','ANO-DE-OBTENCAO-DO-TITULO'], 'count':False, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Graduação', 'tag':'GRADUACAO', 'attr':['NOME-INSTITUICAO','CODIGO-INSTITUICAO','ANO-DE-CONCLUSAO','ANO-DE-OBTENCAO-DO-TITULO'], 'count':False, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Grande área de atuação', 'tag':'AREA-DE-ATUACAO', 'attr':['NOME-GRANDE-AREA-DO-CONHECIMENTO'], 'count':False, 'fnd_blnk':False, 'get_unq':True},
	{'fld_name':'Área de atuação', 'tag':'AREA-DE-ATUACAO', 'attr':['NOME-DA-AREA-DO-CONHECIMENTO'], 'count':False, 'fnd_blnk':False, 'get_unq':True},
	{'fld_name':'Sub-área de atuação', 'tag':'AREA-DE-ATUACAO', 'attr':['NOME-DA-SUB-AREA-DO-CONHECIMENTO'], 'count':False, 'fnd_blnk':False, 'get_unq':True},
	{'fld_name':'Especialidade', 'tag':'AREA-DE-ATUACAO', 'attr':['NOME-DA-ESPECIALIDADE'], 'count':False, 'fnd_blnk':False, 'get_unq':True},
	{'fld_name':'Trabalhos em eventos', 'tag':'TRABALHO-EM-EVENTOS', 'attr':[''], 'count':True, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Artigos publicados', 'tag':'ARTIGO-PUBLICADO', 'attr':[''], 'count':True, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Livros e capítulos', 'tag':'CAPITULO-DE-LIVRO-PUBLICADO', 'attr':[''], 'count':True, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Participação em projetos', 'tag':'PARTICIPACAO-EM-PROJETO', 'attr':[''], 'count':True, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Patentes', 'tag':'PATENTE', 'attr':[''], 'count':True, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Processos ou técnicas', 'tag':'PROCESSOS-OU-TECNICAS', 'attr':[''], 'count':True, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Trabalho técnico', 'tag':'TRABALHO-TECNICO', 'attr':[''], 'count':True, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Orientações (doutorado)', 'tag':'ORIENTACOES-CONCLUIDAS-PARA-DOUTORADO', 'attr':[''], 'count':True, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Orientações (mestrado)', 'tag':'ORIENTACOES-CONCLUIDAS-PARA-MESTRADO', 'attr':[''], 'count':True, 'fnd_blnk':False, 'get_unq':False},
	{'fld_name':'Orientações (outras)', 'tag':'OUTRAS-ORIENTACOES-CONCLUIDAS', 'attr':[''], 'count':True, 'fnd_blnk':False, 'get_unq':False},
	]

app = QApplication(sys.argv)
d = dialogo()
xmls_2_xlsx(d.folder, xml_fields, d.output_file_name)




































# chart_data = [['BarChart', 'bar', 'Formação', 'Formação', '', 'Contagem', [29,29,1,5], [28,28,2,5]],
# 	['BarChart', 'bar', 'Grandes_áreas', 'Grandes áreas de formação', '', 'Contagem', [21,21,1,1000], [20,20,2,1000]],
# 	['PieChart', 'bar', 'Grandes_áreas_perc', 'Grandes áreas de formação', '', 'Contagem', [21,21,1,1000], [20,20,2,1000]],
# 	['BarChart', 'bar', 'Áreas', 'Área de atuação', '', 'Contagem', [26,26,1,1000], [25,25,2,1000]],
# 	['PieChart', 'bar', 'Áreas_perc', 'Área de atuação', '', 'Contagem', [26,26,1,1000], [25,25,2,1000]],
# 	['BarChart', 'bar', 'Produção', 'Produção', '', 'Contagem', [14,14,1,1000], [13,13,2,1000]],
# 	['BarChart', 'bar', 'Produção_relativa', 'Produção/pessoa', '', 'Contagem', [16,16,1,1000], [15,15,2,1000]],
# 	['PieChart', 'bar', 'Italianos_estado', 'Italianos por estado', '', 'Contagem', [8,8,1,30], [6,6,2,30]],
# 	]

# analysis_list = [['Total de italianos', '=COUNTIF(Dados!B:B;"Itália")'],
# 	[],
# 	['País de nascimento', '=SORT(UNIQUE(Dados!B2:B1048576))'],
# 	['Contagem', '=ARRAYFORMULA(COUNTIF(Dados!B2:B1048576,SORT(UNIQUE(Dados!B2:B1048576))))'],
# 	['', ''],
# 	['Estado', '={"AC"; "AL"; "AP"; "AM"; "BA"; "CE"; "DF"; "ES"; "GO"; "MA"; "MT"; "MS"; "MG"; "PA"; "PB"; "PR"; "PE"; "PI"; "RJ"; "RN"; "RS"; "RO"; "RR"; "SC"; "SP"; "SE"; "TO"}'],
# 	['Capital', '={"Rio Branco"; "Maceió"; "Macapá"; "Manaus"; "Salvador"; "Fortaleza"; "Brasília"; "Vitória"; "Goiânia"; "São Luís"; "Cuiabá"; "Campo Grande"; "Belo Horizonte"; "Belém"; "João Pessoa"; "Curitiba"; "Recife"; "Teresina"; "Rio de Janeiro"; "Natal"; "Porto Alegre"; "Porto Velho"; "Boa Vista"; "Florianópolis"; "São Paulo"; "Aracaju"; "Palmas"}'],
# 	['Contagem', '=ARRAYFORMULA(COUNTIF(Italianos!H2:H1048576,filter(F2:F28,F2:F28<>"")))'],
# 	['', ''],
# 	['Produção', '={"Artigos publicados"; "Livros e capítulos"; "Participação em projetos"; "Patentes"; "Processos ou técnicas"; "Trabalho técnico"; "Orientações (doutorado)"; "Orientações (mestrado)"; "Orientações (outras)"}'],
# 	['Contagem absoluta', '={sum(Italianos!S:S); sum(Italianos!T:T); sum(Italianos!U:U); sum(Italianos!V:V); sum(Italianos!W:W); sum(Italianos!X:X); sum(Italianos!Y:Y); sum(Italianos!Z:Z); sum(Italianos!AA:AA)}'],
# 	['Contagem relativa', '={K2/$A$2; K3/$A$2; K4/$A$2; K5/$A$2; K6/$A$2; K7/$A$2; K8/$A$2; K9/$A$2; K10/$A$2}'],
# 	['=J1', '=SORT(J2:K1048576,2,FALSE)'],
# 	['=K1', ''],
# 	['=M1', '=SORT({J2:J1048576,L2:L1048576},2,FALSE)'],
# 	['=L1',''],
# 	['', ''],
# 	['=Italianos!N1', '=sort(unique(transpose(split(join("/",unique(Italianos!N2:N1048576)),"/"))))'],
# 	['Contagem', '=ARRAYFORMULA(COUNTIF(Italianos!$N$2:$N$1048576,ARRAYFORMULA("*" & filter(R2:R1048576,R2:R1048576<>"") & "*")))'],
# 	['=R1', '=SORT(R2:S1048576,2,FALSE)'],
# 	['Contagem', ''],
# 	['', ''],
# 	['=Italianos!O1', '=sort(unique(transpose(split(join("/",unique(Italianos!O2:O1048576)),"/"))))'],
# 	['Contagem', '=ARRAYFORMULA(COUNTIF(Italianos!$O$2:$O1048576,ARRAYFORMULA("*" & filter(W2:W1048576,W2:W1048576<>"") & "*")))'],
# 	['=W1', '=SORT(W2:X1048576,2,FALSE)'],
# 	['Contagem', ''],
# 	['',''],
# 	['Formação', '={Italianos!J1;Italianos!K1;Italianos!L1;Italianos!M1}'],
# 	['Contagem', '={COUNTA(Italianos!J2:J1048576); COUNTA(Italianos!K2:K1048576); COUNTA(Italianos!L2:L1048576); COUNTA(Italianos!M2:M1048576)}'],
# 	['', ''],
# 	['', ''],
# 	['=Italianos!P1', '=sort(unique(transpose(split(join("/",unique(Italianos!P2:P1048576)),"/"))))'],
# 	['Contagem', '=ARRAYFORMULA(COUNTIF(Italianos!$P$2:$P1048576,ARRAYFORMULA("*" & filter(AF2:AF1048576,AF2:AF1048576<>"") & "*")))'],
# 	['=AF1', '=SORT(AF2:AG1048576,2,FALSE)'],
# 	['Contagem', ''],
# 	['', ''],
# 	['=Italianos!Q1', '=sort(unique(transpose(split(join("/",unique(Italianos!Q2:Q1048576)),"/"))))'],
# 	['Contagem', '=ARRAYFORMULA(COUNTIF(Italianos!$Q$2:$Q1048576,ARRAYFORMULA("*" & filter(AK2:AK1048576,AK2:AK1048576<>"") & "*")))'],
# 	['=AK1', '=SORT(AK2:AL1048576,2,FALSE)'],
# 	['Contagem', ''],
# 	]

# País de atuação vs país de graduação - atuação ok, graduação não
# Incluir códigos das instituições (graduação, mestrado e doutorado) - ok
# Conferir origem desses códigos (Google, email ou falar com fuinha)
#    http://di.cnpq.br/di/index.jsp

# Instituição atual onde trabalha é o mais importante. Vínculo atual é importante também - ok
# Incluir contagem de projetos - ok
# Separar brasileiros formados na Itália (graduação etc.)
# EXTRA: Buscar Web of science e Scopus
# https://openpyxl.readthedocs.io/en/stable/usage.html